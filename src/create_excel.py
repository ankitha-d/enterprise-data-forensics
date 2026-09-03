from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


BASE_DIR = Path(__file__).resolve().parent.parent
INPUT_PATH = BASE_DIR / "data" / "processed" / "exception_register.csv"
OUTPUT_DIR = BASE_DIR / "excel"
OUTPUT_PATH = OUTPUT_DIR / "exception_register.xlsx"


def create_workbook():
    print("=" * 60)
    print("ENTERPRISE DATA FORENSICS")
    print("Excel Investigation Workbook")
    print("=" * 60)

    print("\nLoading exception register...")
    df = pd.read_csv(INPUT_PATH)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    print(f"Exception records: {len(df):,}")

    # --------------------------------------------------------
    # Prepare data
    # --------------------------------------------------------

    df["financial_impact"] = pd.to_numeric(
        df["financial_impact"],
        errors="coerce"
    ).fillna(0)

    df["erp_amount"] = pd.to_numeric(
        df["erp_amount"],
        errors="coerce"
    )

    df["payment_amount"] = pd.to_numeric(
        df["payment_amount"],
        errors="coerce"
    )

    df["bank_amount"] = pd.to_numeric(
        df["bank_amount"],
        errors="coerce"
    )

    df["payment_difference"] = pd.to_numeric(
        df["payment_difference"],
        errors="coerce"
    )

    df["bank_difference"] = pd.to_numeric(
        df["bank_difference"],
        errors="coerce"
    )

    # --------------------------------------------------------
    # Workbook
    # --------------------------------------------------------

    wb = Workbook()

    ws_summary = wb.active
    ws_summary.title = "Summary"

    ws_register = wb.create_sheet("Exception Register")

    # --------------------------------------------------------
    # Summary sheet
    # --------------------------------------------------------

    ws_summary["A1"] = "ENTERPRISE DATA FORENSICS"
    ws_summary["A1"].font = Font(size=18, bold=True)

    ws_summary["A2"] = "Exception Investigation Summary"
    ws_summary["A2"].font = Font(size=12, bold=True)

    ws_summary["A4"] = "Metric"
    ws_summary["B4"] = "Value"

    summary_headers = ["Metric", "Value"]

    for cell in ws_summary[4]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(
            "solid",
            fgColor="D9EAF7"
        )

    ws_summary["A5"] = "Total Exceptions"
    ws_summary["B5"] = len(df)

    ws_summary["A6"] = "Financial Exposure"
    ws_summary["B6"] = float(df["financial_impact"].sum())

    ws_summary["A7"] = "Critical Exceptions"
    ws_summary["B7"] = int(
        (df["severity"] == "CRITICAL").sum()
    )

    ws_summary["A8"] = "High Exceptions"
    ws_summary["B8"] = int(
        (df["severity"] == "HIGH").sum()
    )

    ws_summary["A9"] = "Medium Exceptions"
    ws_summary["B9"] = int(
        (df["severity"] == "MEDIUM").sum()
    )

    ws_summary["A11"] = "Exceptions by Type"
    ws_summary["A11"].font = Font(bold=True)

    type_summary = (
        df.groupby("reconciliation_status")
        .agg(
            exception_count=("exception_id", "count"),
            financial_exposure=("financial_impact", "sum")
        )
        .sort_values(
            "financial_exposure",
            ascending=False
        )
    )

    ws_summary["A12"] = "Exception Type"
    ws_summary["B12"] = "Count"
    ws_summary["C12"] = "Financial Exposure"

    for cell in ws_summary[12]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(
            "solid",
            fgColor="D9EAF7"
        )

    row = 13

    for exception_type, values in type_summary.iterrows():
        ws_summary.cell(row, 1, exception_type)
        ws_summary.cell(
            row,
            2,
            int(values["exception_count"])
        )
        ws_summary.cell(
            row,
            3,
            float(values["financial_exposure"])
        )
        row += 1

    # --------------------------------------------------------
    # Exception register
    # --------------------------------------------------------

    columns = list(df.columns)

    for col_idx, column_name in enumerate(columns, start=1):
        cell = ws_register.cell(
            row=1,
            column=col_idx,
            value=column_name
        )

        cell.font = Font(
            bold=True,
            color="FFFFFF"
        )

        cell.fill = PatternFill(
            "solid",
            fgColor="1F4E78"
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    for row_idx, values in enumerate(
        df.itertuples(index=False, name=None),
        start=2
    ):
        for col_idx, value in enumerate(values, start=1):
            ws_register.cell(
                row=row_idx,
                column=col_idx,
                value=value
            )

    # --------------------------------------------------------
    # Excel table
    # --------------------------------------------------------

    last_column = get_column_letter(len(columns))
    last_row = len(df) + 1

    table = Table(
        displayName="ExceptionRegister",
        ref=f"A1:{last_column}{last_row}"
    )

    table_style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )

    table.tableStyleInfo = table_style
    ws_register.add_table(table)

    # --------------------------------------------------------
    # Formatting
    # --------------------------------------------------------

    currency_columns = {
        "financial_impact",
        "erp_amount",
        "payment_amount",
        "payment_difference",
        "bank_amount",
        "bank_difference"
    }

    for col_idx, column_name in enumerate(columns, start=1):
        column_letter = get_column_letter(col_idx)

        for row_idx in range(2, last_row + 1):
            cell = ws_register.cell(
                row=row_idx,
                column=col_idx
            )

            if column_name in currency_columns:
                cell.number_format = '₹#,##0.00'

        max_length = max(
            len(str(column_name)),
            min(
                40,
                max(
                    (
                        len(str(value))
                        for value in df[column_name].head(1000)
                        if pd.notna(value)
                    ),
                    default=0
                )
            )
        )

        ws_register.column_dimensions[
            column_letter
        ].width = max_length + 2

    ws_register.freeze_panes = "A2"
    ws_register.auto_filter.ref = (
        f"A1:{last_column}{last_row}"
    )

    # --------------------------------------------------------
    # Conditional formatting
    # --------------------------------------------------------

    severity_column = columns.index("severity") + 1
    severity_letter = get_column_letter(severity_column)

    ws_register.conditional_formatting.add(
        f"{severity_letter}2:{severity_letter}{last_row}",
        ColorScaleRule(
            start_type="min",
            start_color="FFFFFF",
            mid_type="percentile",
            mid_value=50,
            mid_color="FFF2CC",
            end_type="max",
            end_color="F4CCCC"
        )
    )

    financial_column = columns.index("financial_impact") + 1
    financial_letter = get_column_letter(financial_column)

    ws_register.conditional_formatting.add(
        f"{financial_letter}2:{financial_letter}{last_row}",
        ColorScaleRule(
            start_type="min",
            start_color="FFFFFF",
            mid_type="percentile",
            mid_value=50,
            mid_color="FFF2CC",
            end_type="max",
            end_color="F4CCCC"
        )
    )

    # --------------------------------------------------------
    # Summary formatting
    # --------------------------------------------------------

    ws_summary["B6"].number_format = '₹#,##0.00'

    for cell in ws_summary["A"]:
        cell.alignment = Alignment(
            vertical="center"
        )

    ws_summary.column_dimensions["A"].width = 30
    ws_summary.column_dimensions["B"].width = 22
    ws_summary.column_dimensions["C"].width = 22

    ws_summary.freeze_panes = "A5"

    # --------------------------------------------------------
    # Save
    # --------------------------------------------------------

    wb.save(OUTPUT_PATH)

    print("\nExcel workbook created successfully.")
    print("-" * 60)
    print(f"Output: {OUTPUT_PATH}")
    print(f"Exception records: {len(df):,}")
    print(f"Financial exposure: ₹{df['financial_impact'].sum():,.2f}")


if __name__ == "__main__":
    create_workbook()